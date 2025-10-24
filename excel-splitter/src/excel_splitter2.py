import os
import openpyxl
from openpyxl import Workbook

def extract_first_column(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    first_column_data = [cell.value for cell in sheet['A'] if cell.value is not None]
    return first_column_data

def split_data(data, chunk_size=300):
    for i in range(0, len(data), chunk_size):
        yield data[i:i + chunk_size]

def save_chunks(chunks, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for index, chunk in enumerate(chunks):
        wb = Workbook()
        ws = wb.active
        for row_idx, item in enumerate(chunk, 1):
            ws.cell(row=row_idx, column=1, value=item)
        file_name = os.path.join(output_dir, f'chunk_{index + 1}.xlsx')
        wb.save(file_name)

def main(input_file, output_dir):
    data = extract_first_column(input_file)
    chunks = split_data(data)
    save_chunks(chunks, output_dir)

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Extract first column from XLSX and split into sub-files.')
    parser.add_argument('input_file', type=str, help='Path to the input XLSX file')
    parser.add_argument('output_dir', type=str, help='Directory to save the output files')
    args = parser.parse_args()

    main(args.input_file, args.output_dir)